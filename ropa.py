#!/usr/bin/env python3
"""
RoPA — Registro de Atividades de Tratamento
LGPD compliance

CLI para gestão do Registro de Atividades de Tratamento (RoPA)
conforme LGPD (Lei 13.709/2018).

Uso:
    python ropa.py novo          Criar nova atividade de tratamento
    python ropa.py listar        Listar todas as atividades
    python ropa.py ver <id>      Detalhes de uma atividade
    python ropa.py editar <id>   Editar uma atividade existente
    python ropa.py validar       Validar completude de todos os registros
    python ropa.py exportar      Exportar RoPA (JSON / CSV / XLSX)
    python ropa.py relatorio     Gerar relatório institucional (PDF)
    python ropa.py seed          Popular base com dados de exemplo
"""

import argparse
import csv
import json
import os
import sqlite3
import sys
import textwrap
from datetime import datetime
from pathlib import Path

import cnil_pia_importer
from modelo_ppsi import (
    CAMPOS_VALIDACAO, SITUACOES, CATEGORIAS_DADOS_FCI, JSON_FIELDS,
    migrar_schema, proxima_versao, campos_estruturais, preenchido,
    parse_lista, parse_dict_tipos, parse_estimativa,
    lista_para_texto, dict_tipos_para_texto, dict_estimativa_para_texto,
)

# ── Dependências opcionais ────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
        TableStyle,
    )
    PDF_OK = True
except ImportError:
    PDF_OK = False

# ── Configuração ──────────────────────────────────────────────────────────────
DB_PATH = Path(os.environ.get("ROPA_DB_PATH", Path(__file__).parent / "ropa.db"))
EXPORT_DIR = Path(__file__).parent / "exports"

# Identidade institucional — configurável via env (genérico por padrão)
ORGANIZACAO   = os.environ.get("ROPA_ORGANIZACAO", "Organização")
UNIDADE       = os.environ.get("ROPA_UNIDADE", "Unidade de Proteção de Dados")
ENCARREGADO   = os.environ.get("ROPA_ENCARREGADO", "Encarregado(a) de Proteção de Dados")
NORMAS_RODAPE = os.environ.get(
    "ROPA_NORMAS_REFERENCIA",
    "Documento produzido nos termos da LGPD – Lei 13.709/2018, Art. 37",
)

# Bases legais LGPD Art. 7º + Art. 11 (dados sensíveis)
BASES_LEGAIS = {
    "I":   "Consentimento do titular (Art. 7º, I)",
    "II":  "Obrigação legal ou regulatória (Art. 7º, II)",
    "III": "Execução de políticas públicas (Art. 7º, III)",
    "IV":  "Estudos por órgão de pesquisa (Art. 7º, IV)",
    "V":   "Execução de contrato (Art. 7º, V)",
    "VI":  "Exercício regular de direitos (Art. 7º, VI)",
    "VII": "Proteção da vida ou incolumidade física (Art. 7º, VII)",
    "VIII":"Tutela da saúde (Art. 7º, VIII)",
    "IX":  "Legítimo interesse (Art. 7º, IX)",
    "X":   "Proteção do crédito (Art. 7º, X)",
    "S-I": "Dados sensíveis – Consentimento específico (Art. 11, I)",
    "S-II":"Dados sensíveis – Obrigação legal / exercício de direitos / políticas públicas (Art. 11, II)",
}

# CAMPOS_VALIDACAO (soma=100) e taxonomia FCI-ANPD vêm de modelo_ppsi.py

# ── Banco de dados ────────────────────────────────────────────────────────────

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Cria o schema se ainda não existir e aplica migração PPSI 2.0."""
    with get_conn() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS atividades (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_atividade        TEXT NOT NULL,
            finalidade            TEXT,
            base_legal            TEXT,
            categorias_titulares  TEXT,
            categorias_dados      TEXT,
            dados_sensiveis       INTEGER DEFAULT 0,
            destinatarios         TEXT,
            transferencia_inter   TEXT,
            prazo_retencao        TEXT,
            medidas_seguranca     TEXT,
            unidade_controladora  TEXT,
            sistema_sei           TEXT,
            observacoes           TEXT,
            criado_em             TEXT DEFAULT (datetime('now','localtime')),
            atualizado_em         TEXT DEFAULT (datetime('now','localtime')),
            ativo                 INTEGER DEFAULT 1
        )""")
        conn.execute("""
        CREATE TABLE IF NOT EXISTS historico (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            atividade_id  INTEGER,
            campo         TEXT,
            valor_antigo  TEXT,
            valor_novo    TEXT,
            alterado_em   TEXT DEFAULT (datetime('now','localtime'))
        )""")
        migrar_schema(conn)


# ── Helpers de terminal ───────────────────────────────────────────────────────

VERDE  = "\033[92m"
AMARELO= "\033[93m"
VERMELHO="\033[91m"
AZUL   = "\033[94m"
NEGRITO= "\033[1m"
RESET  = "\033[0m"
CINZA  = "\033[90m"

def banner():
    print(f"""
{AZUL}{NEGRITO}╔══════════════════════════════════════════════════════════╗
║   RoPA         ·  Registro de Atividades de Tratamento   ║
║   LGPD · Lei 13.709/2018                                 ║
╚══════════════════════════════════════════════════════════╝{RESET}
""")


def ok(msg):  print(f"  {VERDE}✔{RESET}  {msg}")
def warn(msg):print(f"  {AMARELO}⚠{RESET}  {msg}")
def erro(msg):print(f"  {VERMELHO}✖{RESET}  {msg}")
def info(msg):print(f"  {AZUL}ℹ{RESET}  {msg}")


def perguntar(campo: str, descricao: str, obrigatorio: bool = False,
              opcoes: dict = None, atual: str = "") -> str:
    """Lê input do usuário com exibição de opções e valor atual."""
    marcador = f"{VERMELHO}*{RESET}" if obrigatorio else " "

    if opcoes:
        print(f"\n  {marcador} {NEGRITO}{descricao}{RESET}")
        for k, v in opcoes.items():
            print(f"      {CINZA}{k:>4}{RESET}  {v}")
        if atual:
            prompt = f"     → [{atual}]: "
        else:
            prompt = f"     → : "
    else:
        if atual:
            prompt = f"\n  {marcador} {NEGRITO}{descricao}{RESET}\n     → [{atual}]: "
        else:
            prompt = f"\n  {marcador} {NEGRITO}{descricao}{RESET}\n     → : "

    while True:
        valor = input(prompt).strip()
        if not valor and atual:
            return atual
        if not valor and obrigatorio:
            warn("Campo obrigatório. Informe um valor.")
            continue
        if opcoes and valor and valor not in opcoes:
            warn(f"Opção inválida. Escolha: {', '.join(opcoes.keys())}")
            continue
        return valor


def sim_nao(pergunta: str, padrao: bool = False) -> bool:
    padrao_str = "S/n" if padrao else "s/N"
    r = input(f"\n  {pergunta} [{padrao_str}]: ").strip().lower()
    if not r:
        return padrao
    return r in ("s", "sim", "y", "yes")


# ── CRUD ──────────────────────────────────────────────────────────────────────

def _json_prompt(atual: dict, campo: str, modo: str, descricao: str,
                 obrigatorio: bool = False) -> str:
    """Prompt para campo JSON, exibindo o texto correspondente como valor atual."""
    val = atual.get(campo)
    if modo == "lista":
        txt = lista_para_texto(val)
    elif modo == "dict":
        txt = dict_tipos_para_texto(val)
    elif modo == "estimativa":
        txt = dict_estimativa_para_texto(val)
    else:
        txt = str(val or "")
    return perguntar(campo, descricao, obrigatorio=obrigatorio, atual=txt)


def _coletar_campos(atual: dict = None) -> dict:
    """Formulário interativo. Se `atual` fornecido, funciona como edição."""
    a = atual or {}

    print(f"\n{NEGRITO}  ── Identificação (Guia 4.1) ────────────────────────────────{RESET}")
    nome         = perguntar("nome_atividade",       "Nome da atividade de tratamento", obrigatorio=True, atual=a.get("nome_atividade",""))
    unidade      = perguntar("unidade_controladora", "Unidade administrativa responsável", atual=a.get("unidade_controladora",""))
    resp_preench = perguntar("responsavel_preenchimento", "Responsável pelo preenchimento (nome, matrícula, data)", atual=a.get("responsavel_preenchimento",""))
    situacao_opts= {k: v for k, v in SITUACOES}
    situacao     = perguntar("situacao", "Situação do registro", opcoes=situacao_opts, atual=a.get("situacao","em_andamento"))
    sistema_sei  = perguntar("sistema_sei",          "Número SEI (processo/documento relacionado)", atual=a.get("sistema_sei",""))

    print(f"\n{NEGRITO}  ── Finalidade e fundamentação (Guia 4.4) ──────────────────{RESET}")
    finalidade   = perguntar("finalidade",   "Finalidade do tratamento", obrigatorio=True, atual=a.get("finalidade",""))
    base_legal   = perguntar("base_legal",   "Base legal", obrigatorio=True, opcoes=BASES_LEGAIS, atual=a.get("base_legal",""))
    prev_norm    = perguntar("previsao_normativa", "Previsão normativa específica (norma e artigos)", atual=a.get("previsao_normativa",""))

    print(f"\n{NEGRITO}  ── Dados pessoais tratados (Guia 4.2) ─────────────────────{RESET}")
    cat_titulares= perguntar("categorias_titulares","Categorias de titulares (ex: servidores, cidadãos)", atual=a.get("categorias_titulares",""))
    tit_estim    = _json_prompt(a, "titulares_estimativa", "estimativa",
                                "Estimativa de titulares por categoria (formato 'categoria: quantidade' por linha)")
    prot_ref     = _json_prompt(a, "titulares_protecao_reforcada", "lista",
                                "Titulares com proteção reforçada (crianças, adolescentes, idosos...; um por linha)")
    cat_dados    = perguntar("categorias_dados","Tipos de dados (resumo: nome, CPF, e-mail...)", atual=a.get("categorias_dados",""))
    tipos_dados  = _json_prompt(a, "tipos_dados", "dict",
                                "Tipos de dados por categoria FCI-ANPD (formato 'Categoria: tipo1; tipo2' por linha)\n  Categorias: " + "; ".join(CATEGORIAS_DADOS_FCI.values()))
    sensiveis    = sim_nao("Envolve dados sensíveis (Art. 5º, II)?", padrao=bool(a.get("dados_sensiveis",0)))
    if sensiveis:
        tipos_sens  = _json_prompt(a, "tipos_dados_sensiveis", "lista",
                                   "Tipos de dados sensíveis (saúde, biometria, religião...; um por linha)")
    else:
        tipos_sens  = ""

    print(f"\n{NEGRITO}  ── Critérios de alto risco (gatilho RIPD – Res. ANPD 2/2022) ──{RESET}")
    tec_emerg   = sim_nao("Usa tecnologias emergentes/ inovadoras?", padrao=bool(a.get("tecnologias_emergentes",0)))
    dec_auto    = sim_nao("Decisões unicamente automatizadas (perfilamento)?", padrao=bool(a.get("decisoes_automatizadas",0)))
    vig_pub     = sim_nao("Vigilância/controle de zonas acessíveis ao público?", padrao=bool(a.get("vigilancia_zonas_publicas",0)))

    print(f"\n{NEGRITO}  ── Ciclo de vida dos dados (Guia 4.3) ─────────────────────{RESET}")
    fluxo        = perguntar("fluxo_tratamento","Descrição do fluxo de tratamento (coleta → uso → eliminação)", atual=a.get("fluxo_tratamento",""))
    origem       = _json_prompt(a, "origem_dados", "lista",
                                "Origem dos dados (formulário, outro sistema, bases públicas...; um por linha)")
    local_arm    = perguntar("local_armazenamento","Local e meio de armazenamento (sistema, base, nuvem, físico)", atual=a.get("local_armazenamento",""))
    prazo        = perguntar("prazo_retencao",   "Período de retenção", atual=a.get("prazo_retencao",""))
    eliminacao   = perguntar("eliminacao_destinacao","Forma de eliminação/destinação final (ex: expurgo, anonimização)", atual=a.get("eliminacao_destinacao",""))
    frequencia   = perguntar("frequencia_tratamento","Frequência do tratamento (contínua, mensal, eventual)", atual=a.get("frequencia_tratamento",""))

    print(f"\n{NEGRITO}  ── Agentes de tratamento (Guia 4.5) ───────────────────────{RESET}")
    controladores= _json_prompt(a, "controladores", "lista",
                                "Controladores (razão social; CNPJ; responsabilidade; um por linha)")
    operadores   = _json_prompt(a, "operadores", "lista",
                                "Operadores (razão social; CNPJ; instrumento contratual; um por linha)")

    print(f"\n{NEGRITO}  ── Compartilhamento e transf. internacional (Guia 4.6) ─────{RESET}")
    destinatarios= perguntar("destinatarios",       "Compartilhamento / destinatários", atual=a.get("destinatarios",""))
    compart      = _json_prompt(a, "compartilhamentos", "lista",
                                "Compartilhamentos (órgão/sistema; finalidade; base; um por linha)")
    transf_inter = _json_prompt(a, "transferencia_internacional", "lista",
                                "Transferência internacional (país; destinatário; tipos; mecanismo; um por linha, 'N/A' se não houver)")
    if not transf_inter:
        transf_inter = '["N/A"]'

    print(f"\n{NEGRITO}  ── Retenção e Segurança ───────────────────────────────────{RESET}")
    medidas      = perguntar("medidas_seguranca","Medidas de segurança adotadas (Art. 46 LGPD)", atual=a.get("medidas_seguranca",""))
    obs          = perguntar("observacoes",      "Observações adicionais", atual=a.get("observacoes",""))

    def _j(v):
        import json as _json
        if isinstance(v, str):
            try:
                return _json.dumps(_json.loads(v), ensure_ascii=False)
            except Exception:
                return _json.dumps(v, ensure_ascii=False)
        return _json.dumps(v, ensure_ascii=False)

    return dict(
        nome_atividade=nome,
        finalidade=finalidade,
        base_legal=base_legal,
        categorias_titulares=cat_titulares,
        categorias_dados=cat_dados,
        dados_sensiveis=int(sensiveis),
        destinatarios=destinatarios,
        transferencia_inter="N/A",
        prazo_retencao=prazo,
        medidas_seguranca=medidas,
        unidade_controladora=unidade,
        sistema_sei=sistema_sei,
        observacoes=obs,
        responsavel_preenchimento=resp_preench,
        situacao=situacao,
        previsao_normativa=prev_norm,
        titulares_estimativa=_j(parse_estimativa(tit_estim)),
        titulares_protecao_reforcada=_j(parse_lista(prot_ref)),
        tipos_dados=_j(parse_dict_tipos(tipos_dados)),
        tipos_dados_sensiveis=_j(parse_lista(tipos_sens)),
        fluxo_tratamento=fluxo,
        origem_dados=_j(parse_lista(origem)),
        local_armazenamento=local_arm,
        eliminacao_destinacao=eliminacao,
        frequencia_tratamento=frequencia,
        controladores=_j(parse_lista(controladores)),
        operadores=_j(parse_lista(operadores)),
        compartilhamentos=_j(parse_lista(compart)),
        transferencia_internacional=_j(parse_lista(transf_inter)),
        tecnologias_emergentes=int(tec_emerg),
        decisoes_automatizadas=int(dec_auto),
        vigilancia_zonas_publicas=int(vig_pub),
    )


def cmd_novo(_args):
    banner()
    print(f"{NEGRITO}  Nova Atividade de Tratamento{RESET}")
    dados = _coletar_campos()
    dados["versao"] = "1.0"
    dados["situacao"] = dados.get("situacao") or "em_andamento"
    with get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO atividades
              (nome_atividade, finalidade, base_legal, categorias_titulares,
               categorias_dados, dados_sensiveis, destinatarios, transferencia_inter,
               prazo_retencao, medidas_seguranca, unidade_controladora, sistema_sei,
               observacoes, responsavel_preenchimento, situacao, versao,
               titulares_estimativa, titulares_protecao_reforcada, tipos_dados,
               tipos_dados_sensiveis, fluxo_tratamento, origem_dados,
               local_armazenamento, eliminacao_destinacao, frequencia_tratamento,
               previsao_normativa, controladores, operadores, compartilhamentos,
               transferencia_internacional, tecnologias_emergentes,
               decisoes_automatizadas, vigilancia_zonas_publicas)
            VALUES
              (:nome_atividade,:finalidade,:base_legal,:categorias_titulares,
               :categorias_dados,:dados_sensiveis,:destinatarios,:transferencia_inter,
               :prazo_retencao,:medidas_seguranca,:unidade_controladora,:sistema_sei,
               :observacoes,:responsavel_preenchimento,:situacao,:versao,
               :titulares_estimativa,:titulares_protecao_reforcada,:tipos_dados,
               :tipos_dados_sensiveis,:fluxo_tratamento,:origem_dados,
               :local_armazenamento,:eliminacao_destinacao,:frequencia_tratamento,
               :previsao_normativa,:controladores,:operadores,:compartilhamentos,
               :transferencia_internacional,:tecnologias_emergentes,
               :decisoes_automatizadas,:vigilancia_zonas_publicas)
        """, dados)
        novo_id = cur.lastrowid
    ok(f"Atividade criada com ID #{novo_id} (versão {dados['versao']})")
    _exibir_pontuacao(dict(dados))


def cmd_listar(_args):
    banner()
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT id, nome_atividade, base_legal, unidade_controladora,
                   dados_sensiveis, atualizado_em
            FROM atividades WHERE ativo=1
            ORDER BY id
        """).fetchall()

    if not rows:
        warn("Nenhuma atividade registrada. Use: python ropa.py novo")
        return

    print(f"  {'ID':>4}  {'Nome':<38}  {'Base':<6}  {'Sensível':^8}  {'Atualizado'}")
    print(f"  {'─'*4}  {'─'*38}  {'─'*6}  {'─'*8}  {'─'*16}")
    for r in rows:
        s = f"{VERMELHO}Sim{RESET}" if r["dados_sensiveis"] else "Não"
        nome_curto = r["nome_atividade"][:37]
        atualizado = r["atualizado_em"][:10] if r["atualizado_em"] else "—"
        print(f"  {r['id']:>4}  {nome_curto:<38}  {(r['base_legal'] or '—'):<6}  {s:^8}  {atualizado}")

    print(f"\n  Total: {len(rows)} atividade(s)  |  Use: python ropa.py ver <id>")


def cmd_ver(args):
    banner()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM atividades WHERE id=? AND ativo=1", (args.id,)
        ).fetchone()

    if not row:
        erro(f"Atividade #{args.id} não encontrada.")
        return

    r = dict(row)
    base_desc = BASES_LEGAIS.get(r["base_legal"], r["base_legal"] or "—")
    sensiveis = f"{VERMELHO}Sim{RESET}" if r["dados_sensiveis"] else "Não"

    print(f"  {NEGRITO}#{r['id']}  {r['nome_atividade']}{RESET}\n")

    def _lbl(campo, val, modo="lista"):
        v = val
        if campo in JSON_FIELDS or modo == "lista":
            v = lista_para_texto(val)
        return v or ""

    linhas = [
        ("Situação",              r.get("situacao") or "em_andamento"),
        ("Versão",                r.get("versao") or "1.0"),
        ("Unidade controladora",  r["unidade_controladora"]),
        ("Responsável pelo preenchimento", r.get("responsavel_preenchimento")),
        ("SEI relacionado",       r["sistema_sei"]),
        ("Finalidade",            r["finalidade"]),
        ("Base legal",            base_desc),
        ("Previsão normativa",    r.get("previsao_normativa")),
        ("Titulares",             r["categorias_titulares"]),
        ("Estimativa titulares",  dict_estimativa_para_texto(r.get("titulares_estimativa"))),
        ("Proteção reforçada",    lista_para_texto(r.get("titulares_protecao_reforcada"))),
        ("Dados pessoais",        r["categorias_dados"]),
        ("Tipos de dados (FCI)",  dict_tipos_para_texto(r.get("tipos_dados"))),
        ("Dados sensíveis",       sensiveis),
        ("Tipos sensíveis",       lista_para_texto(r.get("tipos_dados_sensiveis"))),
        ("Fluxo de tratamento",   r.get("fluxo_tratamento")),
        ("Origem dos dados",      lista_para_texto(r.get("origem_dados"))),
        ("Local armazenamento",   r.get("local_armazenamento")),
        ("Prazo de retenção",     r["prazo_retencao"]),
        ("Eliminação/destinação", r.get("eliminacao_destinacao")),
        ("Frequência",            r.get("frequencia_tratamento")),
        ("Controladores",         lista_para_texto(r.get("controladores"))),
        ("Operadores",            lista_para_texto(r.get("operadores"))),
        ("Compartilhamento",      r["destinatarios"]),
        ("Compartilhamentos",     lista_para_texto(r.get("compartilhamentos"))),
        ("Transf. internacional", r["transferencia_inter"]),
        ("Transf. int. detalhada", lista_para_texto(r.get("transferencia_internacional"))),
        ("Medidas de segurança",  r["medidas_seguranca"]),
        ("Observações",           r["observacoes"]),
        ("Criado em",             r["criado_em"]),
        ("Atualizado em",         r["atualizado_em"]),
    ]
    for label, val in linhas:
        v = val or f"{CINZA}—{RESET}"
        print(f"  {CINZA}{label:<30}{RESET} {v}")

    print()
    _exibir_pontuacao(r)


def cmd_editar(args):
    banner()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM atividades WHERE id=? AND ativo=1", (args.id,)
        ).fetchone()

    if not row:
        erro(f"Atividade #{args.id} não encontrada.")
        return

    atual = dict(row)
    print(f"  {NEGRITO}Editando #{args.id}: {atual['nome_atividade']}{RESET}")
    print(f"  {CINZA}(Enter para manter o valor atual){RESET}\n")

    novos = _coletar_campos(atual=atual)

    with get_conn() as conn:
        # Auditoria campo a campo
        for campo, (_, _peso) in CAMPOS_VALIDACAO.items():
            if novos.get(campo) != atual.get(campo):
                conn.execute("""
                    INSERT INTO historico (atividade_id, campo, valor_antigo, valor_novo)
                    VALUES (?,?,?,?)
                """, (args.id, campo, atual.get(campo), novos.get(campo)))

        # Versionamento semântico (Guia 4.1.6)
        estrutural = any(novos.get(c) != atual.get(c) for c in campos_estruturais())
        nova_versao = proxima_versao(atual.get("versao"), estrutural)
        alterados = [desc for campo, (desc, _) in CAMPOS_VALIDACAO.items()
                     if novos.get(campo) != atual.get(campo)]
        sintese = f"v{nova_versao} – " + (", ".join(alterados[:6]) if alterados else "ajuste menor")
        novos["versao"] = nova_versao
        novos["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        novos["id"] = args.id

        snapshot = json.dumps({**atual, **novos}, ensure_ascii=False, default=str)
        conn.execute("""
            INSERT INTO versoes (atividade_id, versao, sintese, responsavel, snapshot)
            VALUES (?,?,?,?,?)
        """, (args.id, nova_versao, sintese, "CLI", snapshot))

        conn.execute("""
            UPDATE atividades SET
              nome_atividade=:nome_atividade, finalidade=:finalidade,
              base_legal=:base_legal, categorias_titulares=:categorias_titulares,
              categorias_dados=:categorias_dados, dados_sensiveis=:dados_sensiveis,
              destinatarios=:destinatarios, transferencia_inter=:transferencia_inter,
              prazo_retencao=:prazo_retencao, medidas_seguranca=:medidas_seguranca,
              unidade_controladora=:unidade_controladora, sistema_sei=:sistema_sei,
              observacoes=:observacoes, responsavel_preenchimento=:responsavel_preenchimento,
              situacao=:situacao, versao=:versao,
              titulares_estimativa=:titulares_estimativa,
              titulares_protecao_reforcada=:titulares_protecao_reforcada,
              tipos_dados=:tipos_dados, tipos_dados_sensiveis=:tipos_dados_sensiveis,
              fluxo_tratamento=:fluxo_tratamento, origem_dados=:origem_dados,
              local_armazenamento=:local_armazenamento,
              eliminacao_destinacao=:eliminacao_destinacao,
              frequencia_tratamento=:frequencia_tratamento,
              previsao_normativa=:previsao_normativa, controladores=:controladores,
              operadores=:operadores, compartilhamentos=:compartilhamentos,
              transferencia_internacional=:transferencia_internacional,
              tecnologias_emergentes=:tecnologias_emergentes,
              decisoes_automatizadas=:decisoes_automatizadas,
              vigilancia_zonas_publicas=:vigilancia_zonas_publicas,
              atualizado_em=:atualizado_em
            WHERE id=:id
        """, novos)

    ok(f"Atividade #{args.id} atualizada (versão {nova_versao}).")
    _exibir_pontuacao(novos)


# ── Validação ─────────────────────────────────────────────────────────────────

def _pontuacao(atividade: dict) -> tuple[int, list]:
    """Retorna (score 0-100, lista de campos faltando). Alinhado ao conteúdo mínimo do Guia PPSI 2.0."""
    score = 0
    faltando = []
    for campo, (descricao, peso) in CAMPOS_VALIDACAO.items():
        if preenchido(campo, atividade.get(campo)):
            score += peso
        else:
            faltando.append((descricao, peso))
    return score, faltando


def _exibir_pontuacao(atividade: dict):
    score, faltando = _pontuacao(atividade)
    cor = VERDE if score >= 80 else (AMARELO if score >= 50 else VERMELHO)
    barra = "█" * (score // 10) + "░" * (10 - score // 10)
    print(f"\n  Completude:  {cor}{barra}  {score}%{RESET}")
    if faltando:
        print(f"  {AMARELO}Campos incompletos:{RESET}")
        for d, p in faltando:
            print(f"    {CINZA}·{RESET} {d}  {CINZA}(peso {p}%){RESET}")


def cmd_validar(_args):
    banner()
    print(f"  {NEGRITO}Validação de Completude — LGPD Art. 37{RESET}\n")
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM atividades WHERE ativo=1 ORDER BY id"
        ).fetchall()

    if not rows:
        warn("Nenhuma atividade registrada.")
        return

    scores = []
    for row in rows:
        r = dict(row)
        score, faltando = _pontuacao(r)
        scores.append(score)
        cor = VERDE if score >= 80 else (AMARELO if score >= 50 else VERMELHO)
        barra = "█" * (score // 10) + "░" * (10 - score // 10)
        print(f"  #{r['id']:>3}  {r['nome_atividade'][:40]:<40}  {cor}{barra}  {score:3}%{RESET}")
        if faltando and score < 100:
            for d, p in faltando:
                print(f"         {CINZA}↳ {d} ({p}%){RESET}")

    media = sum(scores) / len(scores)
    cor = VERDE if media >= 80 else (AMARELO if media >= 50 else VERMELHO)
    print(f"\n  {'─'*65}")
    print(f"  Média geral: {cor}{media:.1f}%{RESET}  ·  {len(rows)} atividade(s)")

    abaixo = [s for s in scores if s < 80]
    if abaixo:
        warn(f"{len(abaixo)} atividade(s) abaixo de 80% — requerem atenção.")
    else:
        ok("Todas as atividades atingem nível mínimo de completude (≥ 80%).")


# ── Exportação ────────────────────────────────────────────────────────────────

def _todos_registros() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM atividades WHERE ativo=1 ORDER BY id"
        ).fetchall()
    return [dict(r) for r in rows]


def cmd_exportar(args):
    banner()
    registros = _todos_registros()
    if not registros:
        warn("Nenhuma atividade para exportar.")
        return

    EXPORT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    formatos = [f.strip().lower() for f in (args.formato or "json,csv,xlsx").split(",")]

    if "json" in formatos:
        path = EXPORT_DIR / f"ropa_{ts}.json"
        with open(path, "w", encoding="utf-8") as f:
            json.dump(registros, f, ensure_ascii=False, indent=2, default=str)
        ok(f"JSON → {path}")

    if "csv" in formatos:
        path = EXPORT_DIR / f"ropa_{ts}.csv"
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=registros[0].keys())
            writer.writeheader()
            writer.writerows(registros)
        ok(f"CSV  → {path}")

    if "xlsx" in formatos:
        if not XLSX_OK:
            erro("openpyxl não instalado. Execute: pip install openpyxl")
        else:
            path = EXPORT_DIR / f"ropa_{ts}.xlsx"
            _exportar_xlsx(registros, path)
            ok(f"XLSX → {path}")


def _render_valor(campo: str, val) -> str:
    """Renderiza valor para exportação (JSON multivalorado vira texto legível)."""
    if campo == "dados_sensiveis":
        return "Sim" if val else "Não"
    if campo == "base_legal" and val:
        return f"{val} – {BASES_LEGAIS.get(val, val)}"
    if campo in ("titulares_estimativa",):
        return dict_estimativa_para_texto(val).replace("\n", "; ")
    if campo in ("tipos_dados",):
        return dict_tipos_para_texto(val).replace("\n", "; ")
    if campo in JSON_FIELDS:
        return lista_para_texto(val).replace("\n", "; ")
    return str(val or "")


def _exportar_xlsx(registros: list[dict], path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RoPA"

    HEADER_FILL  = PatternFill("solid", fgColor="1F3D7A")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    SENSIVEL_FILL= PatternFill("solid", fgColor="FFE0E0")
    ALT_FILL     = PatternFill("solid", fgColor="F0F4FF")

    colunas_exibir = [
        ("id",                   "ID",                  6),
        ("nome_atividade",       "Atividade",           36),
        ("situacao",             "Situação",            12),
        ("versao",               "Versão",              8),
        ("finalidade",           "Finalidade",          30),
        ("base_legal",           "Base Legal",          8),
        ("unidade_controladora", "Unidade",             20),
        ("responsavel_preenchimento", "Responsável",    22),
        ("previsao_normativa",   "Previsão normativa",  22),
        ("categorias_titulares", "Titulares",           22),
        ("titulares_estimativa", "Estimativa titulares",20),
        ("titulares_protecao_reforcada", "Proteção reforçada", 18),
        ("categorias_dados",     "Dados",               24),
        ("tipos_dados",          "Tipos de dados (FCI)",30),
        ("dados_sensiveis",      "Sensível",            9),
        ("tipos_dados_sensiveis","Tipos sensíveis",     22),
        ("fluxo_tratamento",     "Fluxo",               30),
        ("origem_dados",         "Origem",              22),
        ("local_armazenamento",  "Armazenamento",       22),
        ("prazo_retencao",       "Retenção",            20),
        ("eliminacao_destinacao","Eliminação",          22),
        ("frequencia_tratamento","Frequência",          14),
        ("controladores",        "Controladores",       22),
        ("operadores",           "Operadores",          22),
        ("destinatarios",        "Compartilhamento",    22),
        ("compartilhamentos",    "Compartilhamentos",   24),
        ("transferencia_inter",  "Transf. int.",        14),
        ("transferencia_internacional", "Transf. int. detalhada", 24),
        ("medidas_seguranca",    "Medidas Seg.",        30),
        ("sistema_sei",          "SEI",                 18),
        ("observacoes",          "Observações",         22),
        ("atualizado_em",        "Atualizado",          16),
    ]

    for col_idx, (_, titulo, largura) in enumerate(colunas_exibir, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 30

    for row_idx, reg in enumerate(registros, 2):
        fill = SENSIVEL_FILL if reg.get("dados_sensiveis") else (ALT_FILL if row_idx % 2 == 0 else None)
        for col_idx, (campo, _, _) in enumerate(colunas_exibir, 1):
            val = _render_valor(campo, reg.get(campo, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(path)


# ── Relatório PDF ─────────────────────────────────────────────────────────────

def cmd_relatorio(args):
    banner()
    if not PDF_OK:
        erro("reportlab não instalado. Execute: pip install reportlab")
        return

    registros = _todos_registros()
    if not registros:
        warn("Nenhuma atividade para incluir no relatório.")
        return

    EXPORT_DIR.mkdir(exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    path= EXPORT_DIR / f"relatorio_ropa_{ts}.pdf"

    _gerar_pdf(registros, path)
    ok(f"PDF  → {path}")


def _gerar_pdf(registros: list[dict], path: Path):
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2.5*cm, bottomMargin=2*cm,
    )

    estilos = getSampleStyleSheet()
    cor_azul = colors.HexColor("#1F3D7A")
    cor_cinza= colors.HexColor("#5A5A5A")
    cor_leve = colors.HexColor("#EEF2FF")

    titulo_style = ParagraphStyle("titulo", fontSize=16, fontName="Helvetica-Bold",
                                  textColor=cor_azul, spaceAfter=4, alignment=TA_CENTER)
    sub_style    = ParagraphStyle("sub",    fontSize=10, fontName="Helvetica",
                                  textColor=cor_cinza, spaceAfter=2, alignment=TA_CENTER)
    secao_style  = ParagraphStyle("secao",  fontSize=11, fontName="Helvetica-Bold",
                                  textColor=cor_azul, spaceBefore=12, spaceAfter=6)
    campo_style  = ParagraphStyle("campo",  fontSize=8,  fontName="Helvetica",
                                  textColor=colors.black, leading=11)
    label_style  = ParagraphStyle("label",  fontSize=8,  fontName="Helvetica-Bold",
                                  textColor=cor_cinza, leading=11)
    rodape_style = ParagraphStyle("rodape", fontSize=7,  fontName="Helvetica",
                                  textColor=cor_cinza, alignment=TA_CENTER)

    story = []

    # ── Capa ──
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(ORGANIZACAO, titulo_style))
    story.append(Paragraph(UNIDADE, sub_style))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=cor_azul))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("REGISTRO DE ATIVIDADES DE TRATAMENTO", titulo_style))
    story.append(Paragraph("Conforme LGPD – Lei 13.709/2018, Art. 37", sub_style))
    story.append(Spacer(1, 0.5*cm))

    data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M")
    meta = [
        ["Data de geração:", data_geracao,
         "Encarregado (DPO):", ENCARREGADO],
        ["Total de atividades:", str(len(registros)),
         "Versão:", f"RoPA-{datetime.now().strftime('%Y%m%d')}"],
    ]
    t_meta = Table(meta, colWidths=[4*cm, 6*cm, 4*cm, 5.5*cm])
    t_meta.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0), (-1,-1), 8),
        ("FONTNAME",  (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",  (2,0), (2,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (0,-1), cor_cinza),
        ("TEXTCOLOR", (2,0), (2,-1), cor_cinza),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [cor_leve, colors.white]),
        ("BOX",       (0,0), (-1,-1), 0.5, cor_cinza),
        ("INNERGRID", (0,0), (-1,-1), 0.3, cor_cinza),
        ("TOPPADDING",   (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",  (0,0),(-1,-1), 6),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))

    # ── Sumário de completude ──
    story.append(Paragraph("Resumo de Completude", secao_style))
    scores_info = []
    for reg in registros:
        score, _ = _pontuacao(reg)
        scores_info.append((reg["id"], reg["nome_atividade"], score))

    sum_data = [["ID", "Atividade de Tratamento", "Completude", "Status"]]
    for rid, nome, score in scores_info:
        status = "✔ Completo" if score >= 80 else ("⚠ Parcial" if score >= 50 else "✖ Incompleto")
        cor_st = colors.green if score >= 80 else (colors.orange if score >= 50 else colors.red)
        sum_data.append([str(rid), nome[:55], f"{score}%",
                         Paragraph(f'<font color="{cor_st.hexval()}">{status}</font>', campo_style)])

    t_sum = Table(sum_data, colWidths=[1.2*cm, 10*cm, 2.5*cm, 3*cm])
    t_sum.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), cor_azul),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("ALIGN",        (0,0), (0,-1), "CENTER"),
        ("ALIGN",        (2,0), (2,-1), "CENTER"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, cor_leve]),
        ("BOX",          (0,0), (-1,-1), 0.5, cor_cinza),
        ("INNERGRID",    (0,0), (-1,-1), 0.3, cor_cinza),
        ("TOPPADDING",   (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ("LEFTPADDING",  (0,0), (-1,-1), 5),
    ]))
    story.append(t_sum)

    media_geral = sum(s for _, _, s in scores_info) / len(scores_info)
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"<b>Média de completude: {media_geral:.1f}%</b>  ·  "
        f"{sum(1 for _,_,s in scores_info if s>=80)} registro(s) completo(s) de {len(registros)}",
        ParagraphStyle("media", fontSize=8, fontName="Helvetica", textColor=cor_cinza)
    ))

    # ── Fichas individuais ──
    story.append(PageBreak())
    story.append(Paragraph("Fichas das Atividades de Tratamento", secao_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))

    campos_ficha = [
        ("situacao",              "Situação do registro"),
        ("versao",                "Versão"),
        ("responsavel_preenchimento", "Responsável pelo preenchimento"),
        ("unidade_controladora",  "Unidade controladora"),
        ("sistema_sei",           "Processo SEI relacionado"),
        ("finalidade",            "Finalidade"),
        ("base_legal",            "Base legal (LGPD)"),
        ("previsao_normativa",    "Previsão normativa específica"),
        ("categorias_titulares",  "Categorias de titulares"),
        ("titulares_estimativa",  "Estimativa de titulares"),
        ("titulares_protecao_reforcada", "Titulares com proteção reforçada"),
        ("categorias_dados",      "Dados pessoais envolvidos"),
        ("tipos_dados",           "Tipos de dados (FCI-ANPD)"),
        ("dados_sensiveis",       "Dados sensíveis (Art. 5º, II)"),
        ("tipos_dados_sensiveis", "Tipos de dados sensíveis"),
        ("fluxo_tratamento",      "Fluxo de tratamento"),
        ("origem_dados",          "Origem dos dados"),
        ("local_armazenamento",   "Local e meio de armazenamento"),
        ("prazo_retencao",        "Prazo de retenção"),
        ("eliminacao_destinacao", "Eliminação/destinação final"),
        ("frequencia_tratamento", "Frequência do tratamento"),
        ("controladores",         "Controladores"),
        ("operadores",            "Operadores"),
        ("destinatarios",         "Compartilhamento"),
        ("compartilhamentos",     "Compartilhamentos detalhados"),
        ("transferencia_inter",   "Transferência internacional"),
        ("transferencia_internacional", "Transferência internacional detalhada"),
        ("medidas_seguranca",     "Medidas de segurança (Art. 46)"),
        ("observacoes",           "Observações"),
        ("criado_em",             "Data de criação"),
        ("atualizado_em",         "Última atualização"),
    ]

    for reg in registros:
        story.append(Spacer(1, 0.4*cm))
        score, faltando = _pontuacao(reg)
        cor_score = colors.green if score >= 80 else (colors.orange if score >= 50 else colors.red)

        cabecalho = Table(
            [[Paragraph(f"<b>#{reg['id']}  {reg['nome_atividade']}</b>",
                        ParagraphStyle("cabe", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white)),
              Paragraph(f'<font color="{cor_score.hexval()}"><b>{score}%</b></font>',
                        ParagraphStyle("pct", fontSize=10, fontName="Helvetica-Bold",
                                       textColor=colors.white, alignment=TA_CENTER))]],
            colWidths=[14*cm, 2.5*cm]
        )
        cabecalho.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), cor_azul),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1), 6),
            ("LEFTPADDING", (0,0), (0,-1), 8),
            ("ALIGN",       (1,0), (1,-1), "CENTER"),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(cabecalho)

        ficha_data = []
        for campo, label in campos_ficha:
            val = _render_valor(campo, reg.get(campo, ""))
            val_str = str(val).strip() if val else "—"
            ficha_data.append([
                Paragraph(label, label_style),
                Paragraph(val_str, campo_style),
            ])

        t_ficha = Table(ficha_data, colWidths=[5*cm, 11.5*cm])
        t_ficha.setStyle(TableStyle([
            ("FONTSIZE",     (0,0),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white, cor_leve]),
            ("BOX",          (0,0),(-1,-1), 0.5, cor_cinza),
            ("INNERGRID",    (0,0),(-1,-1), 0.3, cor_cinza),
            ("TOPPADDING",   (0,0),(-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ("LEFTPADDING",  (0,0),(-1,-1), 6),
            ("VALIGN",       (0,0),(-1,-1), "TOP"),
        ]))
        story.append(t_ficha)

    # ── Rodapé ──
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"{ORGANIZACAO}  ·  {UNIDADE}  ·  "
        f"Encarregado: {ENCARREGADO}  ·  Gerado em {data_geracao}",
        rodape_style
    ))
    story.append(Paragraph(
        NORMAS_RODAPE,
        rodape_style
    ))

    doc.build(story)


# ── Seed com dados de exemplo ─────────────────────────────────────────────────

def cmd_importar(args):
    """Import CNIL PIA JSON file."""
    if not args.arquivo:
        err("Informe o caminho do arquivo: python ropa.py importar <arquivo.json>")
        sys.exit(1)

    arquivo_path = Path(args.arquivo)
    if not arquivo_path.exists():
        err(f"Arquivo não encontrado: {arquivo_path}")
        sys.exit(1)

    if not arquivo_path.suffix.lower() == ".json":
        err("Arquivo deve ser JSON (.json)")
        sys.exit(1)

    print()
    print(f"{AZUL}{NEGRITO}Importando PIA CNIL...{RESET}")
    print(f"Arquivo: {arquivo_path}")
    print(f"Estratégia: {args.strategy}\n")

    try:
        inserted, skipped, errors, error_msgs = cnil_pia_importer.import_from_file(
            str(arquivo_path),
            conflict_strategy=args.strategy
        )

        print()
        ok(f"Importação concluída")
        ok(f"Inseridos: {inserted}")
        ok(f"Pulados: {skipped}")

        if errors > 0:
            print(f"\n{VERMELHO}⚠ Erros encontrados:{RESET}")
            for msg in error_msgs[:10]:  # Show first 10 errors
                print(f"  • {msg}")
            if errors > 10:
                print(f"  ... e mais {errors - 10} erros")

        print()
        sys.exit(0 if errors == 0 else 1)

    except Exception as e:
        err(f"Erro ao importar: {str(e)}")
        sys.exit(1)


def cmd_seed(_args):
    banner()
    exemplos = [
        dict(
            nome_atividade="Folha de Pagamento de Colaboradores",
            finalidade="Processamento da remuneração, encargos e benefícios dos colaboradores da Organização",
            base_legal="II",
            categorias_titulares="Colaboradores efetivos, comissionados e contratados",
            categorias_dados="Nome, CPF, matrícula, conta bancária, dependentes, dados previdenciários",
            dados_sensiveis=0,
            destinatarios="Órgão de arrecadação federal, instituição financeira pagadora, sistema de gestão de pessoas",
            transferencia_inter="N/A",
            prazo_retencao="20 anos para documentos trabalhistas; dados previdenciários: permanente",
            medidas_seguranca="Acesso restrito por perfil na área de gestão de pessoas; autenticação forte; canais cifrados com órgãos externos",
            unidade_controladora="Unidade de Gestão de Pessoas",
            sistema_sei="",
            observacoes=""
        ),
        dict(
            nome_atividade="Recrutamento e Seleção de Pessoal",
            finalidade="Gestão de processos seletivos e concursos para ingresso de novos colaboradores",
            base_legal="III",
            categorias_titulares="Candidatos inscritos em processos seletivos",
            categorias_dados="Nome, CPF, e-mail, telefone, escolaridade, histórico profissional, fotografia",
            dados_sensiveis=0,
            destinatarios="Comissão organizadora do certame; sistema oficial de inscrições",
            transferencia_inter="N/A",
            prazo_retencao="5 anos após o encerramento do certame, conforme normas internas",
            medidas_seguranca="Acesso restrito à comissão; ambiente controlado de provas; registros de auditoria",
            unidade_controladora="Unidade de Gestão de Pessoas",
            sistema_sei="",
            observacoes=""
        ),
        dict(
            nome_atividade="Atendimento ao Cidadão (Protocolo e Ouvidoria)",
            finalidade="Registro e tramitação de manifestações, solicitações e pedidos de informação dos cidadãos",
            base_legal="III",
            categorias_titulares="Cidadãos, solicitantes, manifestantes",
            categorias_dados="Nome, CPF, e-mail, telefone, conteúdo da manifestação, dados do atendimento",
            dados_sensiveis=0,
            destinatarios="Áreas internas responsáveis pela resposta; ouvidoria",
            transferencia_inter="N/A",
            prazo_retencao="Prazo legal de guarda de documentos administrativos, conforme normas internas",
            medidas_seguranca="Sistema de protocolo com perfis de acesso; sigilo das manifestações; trilha de auditoria",
            unidade_controladora="Unidade de Atendimento ao Cidadão",
            sistema_sei="",
            observacoes=""
        ),
        dict(
            nome_atividade="Controle de Acesso às Dependências",
            finalidade="Segurança patrimonial e controle de acesso de colaboradores e visitantes às dependências",
            base_legal="IX",
            categorias_titulares="Colaboradores, visitantes, prestadores de serviço",
            categorias_dados="Nome, CPF, horários de entrada e saída, imagem de identificação, dados biométricos",
            dados_sensiveis=1,
            destinatarios="Equipe de segurança institucional; autoridades policiais (incidentes); sem compartilhamento rotineiro",
            transferencia_inter="N/A",
            prazo_retencao="30 dias em sobrescrita contínua; incidentes: até encerramento de apuração",
            medidas_seguranca="Sistema de acesso com registro individual; sala de monitoramento com controle de acesso; sem transmissão externa",
            unidade_controladora="Unidade de Infraestrutura e Segurança",
            sistema_sei="",
            observacoes="RIPD recomendado (Art. 10, §3 LGPD – legítimo interesse + Art. 5º, II – dado sensível biométrico)"
        ),
        dict(
            nome_atividade="Processo Administrativo Disciplinar",
            finalidade="Apuração de irregularidades funcionais de colaboradores, conforme legislação aplicável",
            base_legal="VI",
            categorias_titulares="Colaboradores investigados, testemunhas, denunciantes",
            categorias_dados="Nome, matrícula, histórico funcional, depoimentos, documentos sigilosos",
            dados_sensiveis=0,
            destinatarios="Comissão processante, autoridade máxima, órgão de controle (se cabível)",
            transferencia_inter="N/A",
            prazo_retencao="10 anos após arquivamento; condenações: permanente",
            medidas_seguranca="Processo eletrônico com restrição de acesso; perfis específicos; impressão controlada",
            unidade_controladora="Unidade de Administração",
            sistema_sei="",
            observacoes="Dados de caráter sigiloso – acesso restrito nos termos da legislação de acesso à informação"
        ),
    ]

    with get_conn() as conn:
        for ex in exemplos:
            ex["situacao"] = "concluido"
            ex["versao"] = "1.0"
            conn.execute("""
                INSERT INTO atividades
                  (nome_atividade, finalidade, base_legal, categorias_titulares,
                   categorias_dados, dados_sensiveis, destinatarios, transferencia_inter,
                   prazo_retencao, medidas_seguranca, unidade_controladora, sistema_sei,
                   observacoes, situacao, versao)
                VALUES
                  (:nome_atividade,:finalidade,:base_legal,:categorias_titulares,
                   :categorias_dados,:dados_sensiveis,:destinatarios,:transferencia_inter,
                   :prazo_retencao,:medidas_seguranca,:unidade_controladora,:sistema_sei,
                   :observacoes,:situacao,:versao)
            """, ex)

    ok(f"{len(exemplos)} atividades de exemplo inseridas.")
    info("Execute  python ropa.py listar  para visualizar.")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    init_db()

    parser = argparse.ArgumentParser(
        prog="ropa.py",
        description="RoPA — Registro de Atividades de Tratamento (LGPD)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""
        Exemplos:
          python ropa.py seed                              Popular com dados de exemplo
          python ropa.py novo                              Criar nova atividade
          python ropa.py listar                            Listar todas as atividades
          python ropa.py ver 1                             Ver detalhes da atividade #1
          python ropa.py editar 1                          Editar atividade #1
          python ropa.py validar                           Checar completude de todas
          python ropa.py exportar                          Exportar JSON + CSV + XLSX
          python ropa.py exportar --formato pdf            Exportar apenas PDF
          python ropa.py relatorio                         Gerar relatório PDF institucional
          python ropa.py importar pia.json                 Importar PIA CNIL (skip conflicts)
          python ropa.py importar pia.json --strategy merge Importar PIA com merge
        """)
    )
    sub = parser.add_subparsers(dest="cmd")

    sub.add_parser("novo",     help="Criar nova atividade de tratamento")
    sub.add_parser("listar",   help="Listar todas as atividades")
    sub.add_parser("validar",  help="Validar completude (Art. 37 LGPD)")
    sub.add_parser("relatorio",help="Gerar relatório institucional PDF")
    sub.add_parser("seed",     help="Popular base com dados de exemplo")

    p_ver    = sub.add_parser("ver",    help="Ver detalhes de uma atividade")
    p_ver.add_argument("id", type=int)

    p_editar = sub.add_parser("editar", help="Editar atividade existente")
    p_editar.add_argument("id", type=int)

    p_export = sub.add_parser("exportar", help="Exportar RoPA (json,csv,xlsx)")
    p_export.add_argument("--formato", default="json,csv,xlsx",
                          help="Formatos separados por vírgula (json,csv,xlsx)")

    p_import = sub.add_parser("importar", help="Importar PIA CNIL JSON")
    p_import.add_argument("arquivo", help="Caminho do arquivo .json CNIL PIA")
    p_import.add_argument("--strategy", choices=["skip", "merge", "overwrite"], default="skip",
                          help="Estratégia de conflito (padrão: skip)")

    args = parser.parse_args()

    cmds = {
        "novo":      cmd_novo,
        "listar":    cmd_listar,
        "ver":       cmd_ver,
        "editar":    cmd_editar,
        "validar":   cmd_validar,
        "exportar":  cmd_exportar,
        "relatorio": cmd_relatorio,
        "importar":  cmd_importar,
        "seed":      cmd_seed,
    }

    if args.cmd in cmds:
        cmds[args.cmd](args)
    else:
        banner()
        parser.print_help()


if __name__ == "__main__":
    main()

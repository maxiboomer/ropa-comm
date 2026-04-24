#!/usr/bin/env python3
"""
RoPA — Web App
LGPD compliance
"""

import csv
import io
import json
import os
import sqlite3
from datetime import datetime
from functools import wraps
from pathlib import Path

from authlib.integrations.flask_client import OAuth
from flask import (Flask, Response, flash, redirect, render_template,
                   request, session, url_for)

import cnil_pia_importer

# ── ropa.py shared logic ──────────────────────────────────────────────────────
DB_PATH = Path(os.environ.get("ROPA_DB_PATH", Path(__file__).parent / "ropa.db"))
EXPORT_DIR = Path(os.environ.get("ROPA_DATA_DIR", Path(__file__).parent)) / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

INSTITUICAO = "Instituição"
UNIDADE_RESPONSAVEL = "Unidade de Proteção de Dados"
ENCARREGADA = "Encarregado(a) de Proteção de Dados"

BASES_LEGAIS = {
    "I":    "Consentimento do titular (Art. 7º, I)",
    "II":   "Obrigação legal ou regulatória (Art. 7º, II)",
    "III":  "Execução de políticas públicas (Art. 7º, III)",
    "IV":   "Estudos por órgão de pesquisa (Art. 7º, IV)",
    "V":    "Execução de contrato (Art. 7º, V)",
    "VI":   "Exercício regular de direitos (Art. 7º, VI)",
    "VII":  "Proteção da vida ou incolumidade física (Art. 7º, VII)",
    "VIII": "Tutela da saúde (Art. 7º, VIII)",
    "IX":   "Legítimo interesse (Art. 7º, IX)",
    "X":    "Proteção do crédito (Art. 7º, X)",
    "S-I":  "Dados sensíveis – Consentimento específico (Art. 11, I)",
    "S-II": "Dados sensíveis – Obrigação legal / exercício de direitos / políticas públicas (Art. 11, II)",
}

CAMPOS_VALIDACAO = {
    "nome_atividade":       ("Nome da atividade de tratamento", 15),
    "finalidade":           ("Finalidade do tratamento", 15),
    "base_legal":           ("Base legal (Art. 7º LGPD)", 15),
    "categorias_titulares": ("Categorias de titulares", 10),
    "categorias_dados":     ("Categorias de dados pessoais", 10),
    "destinatarios":        ("Destinatários / compartilhamento", 10),
    "prazo_retencao":       ("Prazo de retenção", 10),
    "medidas_seguranca":    ("Medidas de segurança (Art. 46)", 10),
    "unidade_controladora": ("Unidade controladora / responsável", 5),
}

EXEMPLOS = [
    dict(
        nome_atividade="Cadastro de Mesários Voluntários",
        finalidade="Recrutamento e gestão de mesários para eleições, conforme Lei 9.504/1997",
        base_legal="III",
        categorias_titulares="Eleitores voluntários cadastrados no sistema Mesário",
        categorias_dados="Nome completo, CPF, título de eleitor, e-mail, telefone, endereço",
        dados_sensiveis=0,
        destinatarios="Unidade de TI, cartórios eleitorais, TSE (integração sistêmica)",
        transferencia_inter="N/A",
        prazo_retencao="5 anos após o pleito, conforme Res. TSE 23.222/2010",
        medidas_seguranca="Acesso restrito por perfil no ELO, autenticação GOV.BR, logs de auditoria",
        unidade_controladora="Unidade de TI",
        sistema_sei="SEI 0006491-21.2026.6.16.8000",
        observacoes="PAD 004717/2022 – sistema em processo de modernização",
    ),
    dict(
        nome_atividade="Registro de Candidaturas (CAND)",
        finalidade="Processamento de pedidos de registro de candidatura para pleitos eleitorais",
        base_legal="III",
        categorias_titulares="Candidatos, vices, suplentes registrados na Justiça Eleitoral",
        categorias_dados="Nome, CPF, título eleitoral, filiação partidária, bens declarados, foto",
        dados_sensiveis=0,
        destinatarios="TSE (DivulgaCand), partidos políticos, imprensa (dados públicos)",
        transferencia_inter="N/A",
        prazo_retencao="Indeterminado – registro histórico eleitoral permanente",
        medidas_seguranca="Sistema CAND/TSE com controle de acesso por OAB/partido; dados públicos via DivulgaCand",
        unidade_controladora="SECJUD – Secretaria Judiciária e de Gestão da Informação",
        sistema_sei="",
        observacoes="",
    ),
    dict(
        nome_atividade="Folha de Pagamento de Servidores",
        finalidade="Processamento da remuneração, encargos e benefícios dos servidores",
        base_legal="II",
        categorias_titulares="Servidores efetivos, comissionados e requisitados",
        categorias_dados="Nome, CPF, matrícula SIAPE, conta bancária, dependentes, dados previdenciários",
        dados_sensiveis=0,
        destinatarios="Receita Federal, PSSS, SIAPE/Ministério da Gestão, CEF",
        transferencia_inter="N/A",
        prazo_retencao="20 anos conforme Res. TSE e TCU; documentos previdenciários: permanente",
        medidas_seguranca="SIAPE com autenticação gov.br; acesso restrito à SECGP; canais cifrados com Receita Federal",
        unidade_controladora="SECGP – Seção de Gestão de Pessoas",
        sistema_sei="",
        observacoes="",
    ),
    dict(
        nome_atividade="Processo Administrativo Disciplinar (PAD)",
        finalidade="Apuração de irregularidades funcionais de servidores, conforme Lei 8.112/1990",
        base_legal="VI",
        categorias_titulares="Servidores investigados, testemunhas, denunciantes",
        categorias_dados="Nome, matrícula, histórico funcional, depoimentos, documentos sigilosos",
        dados_sensiveis=0,
        destinatarios="Comissão processante, Presidência, CGU (se cabível)",
        transferencia_inter="N/A",
        prazo_retencao="10 anos após arquivamento; condenações: permanente",
        medidas_seguranca="Processo SEI com restrição de acesso; perfis específicos; impressão controlada",
        unidade_controladora="SECAD – Secretaria de Administração",
        sistema_sei="",
        observacoes="Dados de caráter sigiloso – acesso restrito nos termos da LAI",
    ),
    dict(
        nome_atividade="Monitoramento por Câmeras (CFTV)",
        finalidade="Segurança patrimonial e controle de acesso às dependências da Instituição",
        base_legal="IX",
        categorias_titulares="Servidores, visitantes, prestadores de serviço",
        categorias_dados="Imagens de vídeo com identificação facial incidental",
        dados_sensiveis=1,
        destinatarios="Segurança institucional; Polícia Federal (incidentes); sem compartilhamento rotineiro",
        transferencia_inter="N/A",
        prazo_retencao="30 dias em sobrescrita contínua; incidentes: até encerramento de apuração",
        medidas_seguranca="DVR com acesso físico restrito; sala de monitoramento com controle de acesso; sem transmissão externa",
        unidade_controladora="ASSEG – Seção de Infraestrutura e Segurança",
        sistema_sei="",
        observacoes="RIPD recomendado (Art. 10, §3 LGPD – legítimo interesse + Art. 5º, II – dado sensível biométrico incidental)",
    ),
]


# ── DB helpers ────────────────────────────────────────────────────────────────

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_conn() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS atividades (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_atividade        TEXT NOT NULL,
            finalidade            TEXT,
            base_legal            TEXT,
            categorias_titulares  TEXT,
            categorias_dados      TEXT,
            dados_sensiveis       INTEGER DEFAULT 0,
            destinatarios         TEXT,
            transferencia_inter   TEXT,
            prazo_retencao        TEXT,
            medidas_seguranca     TEXT,
            unidade_controladora  TEXT,
            sistema_sei           TEXT,
            observacoes           TEXT,
            criado_em             TEXT DEFAULT (datetime('now','localtime')),
            atualizado_em         TEXT DEFAULT (datetime('now','localtime')),
            ativo                 INTEGER DEFAULT 1
        )""")
        conn.execute("""
        CREATE TABLE IF NOT EXISTS historico (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            atividade_id  INTEGER,
            campo         TEXT,
            valor_antigo  TEXT,
            valor_novo    TEXT,
            alterado_em   TEXT DEFAULT (datetime('now','localtime'))
        )""")


def todos_registros() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM atividades WHERE ativo=1 ORDER BY id"
        ).fetchall()
    return [dict(r) for r in rows]


def get_atividade(atividade_id: int):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM atividades WHERE id=? AND ativo=1", (atividade_id,)
        ).fetchone()
    return dict(row) if row else None


def pontuacao(atividade: dict) -> tuple[int, list]:
    score = 0
    faltando = []
    for campo, (descricao, peso) in CAMPOS_VALIDACAO.items():
        val = atividade.get(campo)
        if val and str(val).strip() and str(val).strip().upper() not in ("N/A", "NENHUM", "—"):
            score += peso
        else:
            faltando.append((descricao, peso))
    return score, faltando


def score_class(score: int) -> str:
    if score >= 80:
        return "success"
    if score >= 50:
        return "warning"
    return "danger"


# ── Flask app ─────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "ropa-dev-only-change-in-prod")

# ── Keycloak OIDC ─────────────────────────────────────────────────────────────

KEYCLOAK_MOCK_MODE = os.environ.get("KEYCLOAK_MOCK", "").lower() in ("1", "true", "yes")

if KEYCLOAK_MOCK_MODE:
    # Mock: endpoints rodam na mesma porta via blueprint
    from keycloak_blueprint import register_mock_keycloak
    ROPA_BASE_URL = os.environ.get("ROPA_BASE_URL", "http://127.0.0.1:8000")
    register_mock_keycloak(app, base_url=ROPA_BASE_URL)
    KEYCLOAK_URL = f"{ROPA_BASE_URL}/mock-kc"
else:
    KEYCLOAK_URL = os.environ.get("KEYCLOAK_URL", "http://localhost:8080")

KEYCLOAK_REALM = os.environ.get("KEYCLOAK_REALM", "ropa")
KEYCLOAK_CLIENT_ID = os.environ.get("KEYCLOAK_CLIENT_ID", "ropa-web")
KEYCLOAK_CLIENT_SECRET = os.environ.get("KEYCLOAK_CLIENT_SECRET", "")

OIDC_BASE = f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}"
OIDC_DISCOVERY = f"{OIDC_BASE}/.well-known/openid-configuration"

oauth = OAuth(app)
oauth.register(
    name="keycloak",
    client_id=KEYCLOAK_CLIENT_ID,
    client_secret=KEYCLOAK_CLIENT_SECRET,
    server_metadata_url=OIDC_DISCOVERY,
    client_kwargs={"scope": "openid email profile"},
)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated


app.jinja_env.globals.update(
    BASES_LEGAIS=BASES_LEGAIS,
    CAMPOS_VALIDACAO=CAMPOS_VALIDACAO,
    INSTITUICAO=INSTITUICAO,
    UNIDADE_RESPONSAVEL=UNIDADE_RESPONSAVEL,
    ENCARREGADA=ENCARREGADA,
    pontuacao=pontuacao,
    score_class=score_class,
    now=datetime.now,
)


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/login")
def login():
    if "user" in session:
        return redirect(url_for("index"))
    error = request.args.get("error")
    return render_template("login.html", error=error)


@app.route("/login/keycloak")
def login_keycloak():
    redirect_uri = url_for("auth_callback", _external=True)
    if KEYCLOAK_MOCK_MODE:
        # Constrói a URL de autorização no host do próprio request (não no
        # ISSUER cacheado), para que o browser consiga alcançar o mock-kc
        # tanto via 127.0.0.1 quanto via URL pública (Cloudflare Tunnel).
        import secrets as _secrets
        import urllib.parse as _up
        state = _secrets.token_urlsafe(16)
        nonce = _secrets.token_urlsafe(16)
        session["oauth_state"] = state
        auth_url = f"{request.host_url.rstrip('/')}/mock-kc/realms/{KEYCLOAK_REALM}/protocol/openid-connect/auth"
        params = _up.urlencode({
            "response_type": "code",
            "client_id": KEYCLOAK_CLIENT_ID,
            "redirect_uri": redirect_uri,
            "scope": "openid email profile",
            "state": state,
            "nonce": nonce,
        })
        return redirect(f"{auth_url}?{params}")
    try:
        return oauth.keycloak.authorize_redirect(redirect_uri)
    except Exception:
        return redirect(url_for("login", error="Servidor Keycloak indisponível. Contate a Unidade de TI."))


@app.route("/auth/callback")
def auth_callback():
    import requests as req

    if KEYCLOAK_MOCK_MODE:
        # Mock: trocar o code manualmente (sem verificação JWT)
        code = request.args.get("code", "")
        if not code:
            return redirect(url_for("login", error="Falha na autenticação."))
        try:
            resp = req.post(f"{OIDC_BASE}/protocol/openid-connect/token", data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": url_for("auth_callback", _external=True),
                "client_id": KEYCLOAK_CLIENT_ID,
                "client_secret": KEYCLOAK_CLIENT_SECRET,
            })
            token = resp.json()
            if "error" in token:
                return redirect(url_for("login", error="Código expirado. Tente novamente."))
            ui_resp = req.get(f"{OIDC_BASE}/protocol/openid-connect/userinfo",
                              headers={"Authorization": f"Bearer {token['access_token']}"})
            userinfo = ui_resp.json()
        except Exception:
            return redirect(url_for("login", error="Falha na autenticação."))
    else:
        # Produção: fluxo OIDC padrão via authlib
        try:
            token = oauth.keycloak.authorize_access_token()
            userinfo = token.get("userinfo", {})
            if not userinfo:
                userinfo = oauth.keycloak.userinfo()
        except Exception:
            return redirect(url_for("login", error="Falha na autenticação. Tente novamente."))

    session["user"] = {
        "sub": userinfo.get("sub", ""),
        "name": userinfo.get("name", userinfo.get("preferred_username", "Usuário")),
        "email": userinfo.get("email", ""),
        "username": userinfo.get("preferred_username", ""),
        "roles": userinfo.get("realm_access", {}).get("roles", []),
    }
    session["id_token"] = token.get("id_token", "")

    flash(f"Bem-vindo(a), {session['user']['name']}!", "success")
    next_url = request.args.get("next", url_for("index"))
    return redirect(next_url)


@app.route("/logout")
def logout():
    id_token = session.pop("id_token", "")
    session.pop("user", None)
    session.clear()

    post_logout_uri = url_for("login", _external=True)
    keycloak_logout = (
        f"{OIDC_BASE}/protocol/openid-connect/logout"
        f"?id_token_hint={id_token}"
        f"&post_logout_redirect_uri={post_logout_uri}"
    )
    return redirect(keycloak_logout)


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    registros = todos_registros()
    scores = [pontuacao(r)[0] for r in registros]
    media = sum(scores) / len(scores) if scores else 0
    completos = sum(1 for s in scores if s >= 80)
    sensiveis = sum(1 for r in registros if r.get("dados_sensiveis"))
    return render_template(
        "index.html",
        registros=registros,
        scores=scores,
        media=media,
        completos=completos,
        sensiveis=sensiveis,
    )


@app.route("/atividades")
@login_required
def listar():
    registros = todos_registros()
    return render_template("listar.html", registros=registros)


@app.route("/atividades/nova", methods=["GET", "POST"])
@login_required
def nova():
    if request.method == "POST":
        dados = _form_to_dict(request.form)
        with get_conn() as conn:
            cur = conn.execute("""
                INSERT INTO atividades
                  (nome_atividade, finalidade, base_legal, categorias_titulares,
                   categorias_dados, dados_sensiveis, destinatarios, transferencia_inter,
                   prazo_retencao, medidas_seguranca, unidade_controladora, sistema_sei,
                   observacoes)
                VALUES
                  (:nome_atividade,:finalidade,:base_legal,:categorias_titulares,
                   :categorias_dados,:dados_sensiveis,:destinatarios,:transferencia_inter,
                   :prazo_retencao,:medidas_seguranca,:unidade_controladora,:sistema_sei,
                   :observacoes)
            """, dados)
            novo_id = cur.lastrowid
        flash(f"Atividade #{novo_id} criada com sucesso.", "success")
        return redirect(url_for("ver", atividade_id=novo_id))
    return render_template("form.html", atividade=None, titulo="Nova Atividade")


@app.route("/atividades/<int:atividade_id>")
@login_required
def ver(atividade_id):
    atividade = get_atividade(atividade_id)
    if not atividade:
        flash(f"Atividade #{atividade_id} não encontrada.", "danger")
        return redirect(url_for("listar"))
    score, faltando = pontuacao(atividade)
    historico = []
    with get_conn() as conn:
        historico = conn.execute(
            "SELECT * FROM historico WHERE atividade_id=? ORDER BY alterado_em DESC LIMIT 20",
            (atividade_id,)
        ).fetchall()
    return render_template(
        "ver.html",
        atividade=atividade,
        score=score,
        faltando=faltando,
        historico=historico,
        base_desc=BASES_LEGAIS.get(atividade.get("base_legal", ""), "—"),
    )


@app.route("/atividades/<int:atividade_id>/editar", methods=["GET", "POST"])
@login_required
def editar(atividade_id):
    atividade = get_atividade(atividade_id)
    if not atividade:
        flash(f"Atividade #{atividade_id} não encontrada.", "danger")
        return redirect(url_for("listar"))

    if request.method == "POST":
        novos = _form_to_dict(request.form)
        with get_conn() as conn:
            for campo in CAMPOS_VALIDACAO:
                if novos.get(campo) != atividade.get(campo):
                    conn.execute("""
                        INSERT INTO historico (atividade_id, campo, valor_antigo, valor_novo)
                        VALUES (?,?,?,?)
                    """, (atividade_id, campo, atividade.get(campo), novos.get(campo)))
            novos["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            novos["id"] = atividade_id
            conn.execute("""
                UPDATE atividades SET
                  nome_atividade=:nome_atividade, finalidade=:finalidade,
                  base_legal=:base_legal, categorias_titulares=:categorias_titulares,
                  categorias_dados=:categorias_dados, dados_sensiveis=:dados_sensiveis,
                  destinatarios=:destinatarios, transferencia_inter=:transferencia_inter,
                  prazo_retencao=:prazo_retencao, medidas_seguranca=:medidas_seguranca,
                  unidade_controladora=:unidade_controladora, sistema_sei=:sistema_sei,
                  observacoes=:observacoes, atualizado_em=:atualizado_em
                WHERE id=:id
            """, novos)
        flash(f"Atividade #{atividade_id} atualizada.", "success")
        return redirect(url_for("ver", atividade_id=atividade_id))

    return render_template(
        "form.html",
        atividade=atividade,
        titulo=f"Editar Atividade #{atividade_id}",
    )


@app.route("/atividades/<int:atividade_id>/excluir", methods=["POST"])
@login_required
def excluir(atividade_id):
    with get_conn() as conn:
        conn.execute("UPDATE atividades SET ativo=0 WHERE id=?", (atividade_id,))
    flash(f"Atividade #{atividade_id} removida.", "warning")
    return redirect(url_for("listar"))


@app.route("/validar")
@login_required
def validar():
    registros = todos_registros()
    resultados = []
    for r in registros:
        score, faltando = pontuacao(r)
        resultados.append({"atividade": r, "score": score, "faltando": faltando})
    scores = [x["score"] for x in resultados]
    media = sum(scores) / len(scores) if scores else 0
    return render_template("validar.html", resultados=resultados, media=media)


@app.route("/importar", methods=["GET", "POST"])
@login_required
def importar():
    if request.method == "GET":
        return render_template("importar.html")

    # POST: Handle file upload
    if "file" not in request.files:
        flash("Nenhum arquivo enviado.", "danger")
        return redirect(url_for("importar"))

    file = request.files["file"]
    if file.filename == "":
        flash("Arquivo vazio.", "danger")
        return redirect(url_for("importar"))

    # Validate file type
    if not file.filename.endswith(".json"):
        flash("Apenas arquivos .json são aceitos.", "danger")
        return redirect(url_for("importar"))

    # Validate file size (max 10MB)
    file.seek(0, 2)  # Seek to end
    size = file.tell()
    file.seek(0)  # Reset to start
    if size > 10 * 1024 * 1024:
        flash("Arquivo muito grande (máximo 10MB).", "danger")
        return redirect(url_for("importar"))

    # Get conflict resolution strategy
    strategy = request.form.get("conflict_strategy", "skip")
    if strategy not in ("skip", "merge", "overwrite"):
        strategy = "skip"

    try:
        # Read and import
        content = file.read().decode("utf-8")
        imported, skipped, errors, error_msgs = cnil_pia_importer.import_from_content(
            content=content,
            conflict_strategy=strategy
        )

        if errors > 0:
            flash(
                f"Importação com problemas: {imported} inseridos, {skipped} pulados, {errors} erros.",
                "warning"
            )
            for msg in error_msgs[:5]:  # Show first 5 errors
                flash(f"  • {msg}", "info")
        else:
            flash(
                f"✓ Importação bem-sucedida: {imported} atividades importadas, {skipped} puladas.",
                "success"
            )

        return redirect(url_for("listar"))

    except Exception as e:
        flash(f"Erro ao importar: {str(e)}", "danger")
        return redirect(url_for("importar"))


@app.route("/exportar")
@login_required
def exportar():
    fmt = request.args.get("formato", "json")
    registros = todos_registros()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if fmt == "json":
        data = json.dumps(registros, ensure_ascii=False, indent=2, default=str)
        return Response(
            data,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename=ropa_{ts}.json"},
        )

    if fmt == "csv":
        output = io.StringIO()
        if registros:
            writer = csv.DictWriter(output, fieldnames=registros[0].keys())
            writer.writeheader()
            writer.writerows(registros)
        return Response(
            "\ufeff" + output.getvalue(),
            mimetype="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f"attachment; filename=ropa_{ts}.csv"},
        )

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            flash("openpyxl não instalado.", "danger")
            return redirect(url_for("index"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RoPA"

        HEADER_FILL = PatternFill("solid", fgColor="1F3D7A")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
        SENSIVEL_FILL = PatternFill("solid", fgColor="FFE0E0")
        ALT_FILL = PatternFill("solid", fgColor="F0F4FF")

        colunas = [
            ("id", "ID", 6),
            ("nome_atividade", "Atividade", 40),
            ("finalidade", "Finalidade", 30),
            ("base_legal", "Base Legal", 8),
            ("categorias_titulares", "Titulares", 25),
            ("categorias_dados", "Dados", 25),
            ("dados_sensiveis", "Sensível", 9),
            ("destinatarios", "Destinatários", 25),
            ("prazo_retencao", "Retenção", 20),
            ("medidas_seguranca", "Medidas Seg.", 30),
            ("unidade_controladora", "Unidade", 20),
            ("sistema_sei", "SEI", 18),
            ("atualizado_em", "Atualizado", 16),
        ]

        for col_idx, (_, titulo, largura) in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = largura
        ws.row_dimensions[1].height = 30

        for row_idx, reg in enumerate(registros, 2):
            fill = SENSIVEL_FILL if reg.get("dados_sensiveis") else (ALT_FILL if row_idx % 2 == 0 else None)
            for col_idx, (campo, _, _) in enumerate(colunas, 1):
                val = reg.get(campo, "")
                if campo == "dados_sensiveis":
                    val = "Sim" if val else "Não"
                elif campo == "base_legal":
                    val = f"{val} – {BASES_LEGAIS.get(val, '')}" if val else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    cell.fill = fill
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=ropa_{ts}.xlsx"},
        )

    if fmt == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (HRFlowable, PageBreak, Paragraph,
                                            SimpleDocTemplate, Spacer, Table,
                                            TableStyle)
        except ImportError:
            flash("reportlab não instalado.", "danger")
            return redirect(url_for("index"))

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm,
            topMargin=2.5 * cm, bottomMargin=2 * cm,
        )

        cor_azul = colors.HexColor("#1F3D7A")
        cor_cinza = colors.HexColor("#5A5A5A")
        cor_leve = colors.HexColor("#EEF2FF")

        titulo_style = ParagraphStyle("titulo", fontSize=16, fontName="Helvetica-Bold",
                                      textColor=cor_azul, spaceAfter=4, alignment=TA_CENTER)
        sub_style = ParagraphStyle("sub", fontSize=10, fontName="Helvetica",
                                   textColor=cor_cinza, spaceAfter=2, alignment=TA_CENTER)
        secao_style = ParagraphStyle("secao", fontSize=11, fontName="Helvetica-Bold",
                                     textColor=cor_azul, spaceBefore=12, spaceAfter=6)
        campo_style = ParagraphStyle("campo", fontSize=8, fontName="Helvetica",
                                     textColor=colors.black, leading=11)
        label_style = ParagraphStyle("label", fontSize=8, fontName="Helvetica-Bold",
                                     textColor=cor_cinza, leading=11)
        rodape_style = ParagraphStyle("rodape", fontSize=7, fontName="Helvetica",
                                      textColor=cor_cinza, alignment=TA_CENTER)

        story = []
        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph(INSTITUICAO, titulo_style))
        story.append(Paragraph(UNIDADE_RESPONSAVEL, sub_style))
        story.append(Spacer(1, 0.3 * cm))
        story.append(HRFlowable(width="100%", thickness=2, color=cor_azul))
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph("REGISTRO DE ATIVIDADES DE TRATAMENTO", titulo_style))
        story.append(Paragraph("Conforme LGPD – Lei 13.709/2018, Art. 37", sub_style))
        story.append(Spacer(1, 0.5 * cm))

        data_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M")
        meta = [
            ["Data de geração:", data_geracao, "Encarregada (DPO):", ENCARREGADA],
            ["Total de atividades:", str(len(registros)), "Versão:", f"RoPA-{datetime.now().strftime('%Y%m%d')}"],
        ]
        t_meta = Table(meta, colWidths=[4 * cm, 6 * cm, 4 * cm, 5.5 * cm])
        t_meta.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [cor_leve, colors.white]),
            ("BOX", (0, 0), (-1, -1), 0.5, cor_cinza),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, cor_cinza),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 0.5 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))

        story.append(Paragraph("Resumo de Completude", secao_style))
        scores_info = [(r["id"], r["nome_atividade"], pontuacao(r)[0]) for r in registros]
        sum_data = [["ID", "Atividade de Tratamento", "Completude", "Status"]]
        for rid, nome, score in scores_info:
            status = "Completo" if score >= 80 else ("Parcial" if score >= 50 else "Incompleto")
            cor_st = colors.green if score >= 80 else (colors.orange if score >= 50 else colors.red)
            sum_data.append([str(rid), nome[:55], f"{score}%",
                             Paragraph(f'<font color="{cor_st.hexval()}">{status}</font>', campo_style)])
        t_sum = Table(sum_data, colWidths=[1.2 * cm, 10 * cm, 2.5 * cm, 3 * cm])
        t_sum.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), cor_azul),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, cor_leve]),
            ("BOX", (0, 0), (-1, -1), 0.5, cor_cinza),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, cor_cinza),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t_sum)

        media_geral = sum(s for _, _, s in scores_info) / len(scores_info) if scores_info else 0
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(
            f"<b>Média de completude: {media_geral:.1f}%</b>  ·  "
            f"{sum(1 for _, _, s in scores_info if s >= 80)} registro(s) completo(s) de {len(registros)}",
            ParagraphStyle("media", fontSize=8, fontName="Helvetica", textColor=cor_cinza),
        ))

        story.append(PageBreak())
        story.append(Paragraph("Fichas das Atividades de Tratamento", secao_style))
        story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))

        campos_ficha = [
            ("finalidade", "Finalidade"),
            ("base_legal", "Base legal (LGPD)"),
            ("categorias_titulares", "Categorias de titulares"),
            ("categorias_dados", "Dados pessoais envolvidos"),
            ("dados_sensiveis", "Dados sensíveis (Art. 5º, II)"),
            ("destinatarios", "Destinatários / compartilhamento"),
            ("transferencia_inter", "Transferência internacional"),
            ("prazo_retencao", "Prazo de retenção"),
            ("medidas_seguranca", "Medidas de segurança (Art. 46)"),
            ("unidade_controladora", "Unidade controladora"),
            ("sistema_sei", "Processo SEI relacionado"),
            ("observacoes", "Observações"),
            ("criado_em", "Data de criação"),
            ("atualizado_em", "Última atualização"),
        ]

        for reg in registros:
            story.append(Spacer(1, 0.4 * cm))
            score, _ = pontuacao(reg)
            cor_score = colors.green if score >= 80 else (colors.orange if score >= 50 else colors.red)
            cabecalho = Table(
                [[Paragraph(f"<b>#{reg['id']}  {reg['nome_atividade']}</b>",
                            ParagraphStyle("cabe", fontSize=10, fontName="Helvetica-Bold", textColor=colors.white)),
                  Paragraph(f'<font color="{cor_score.hexval()}"><b>{score}%</b></font>',
                            ParagraphStyle("pct", fontSize=10, fontName="Helvetica-Bold",
                                           textColor=colors.white, alignment=TA_CENTER))]],
                colWidths=[14 * cm, 2.5 * cm],
            )
            cabecalho.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), cor_azul),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (0, -1), 8),
                ("ALIGN", (1, 0), (1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(cabecalho)

            ficha_data = []
            for campo, label in campos_ficha:
                val = reg.get(campo, "")
                if campo == "dados_sensiveis":
                    val = "Sim" if val else "Não"
                elif campo == "base_legal" and val:
                    val = f"{val} – {BASES_LEGAIS.get(val, val)}"
                val_str = str(val).strip() if val else "—"
                ficha_data.append([Paragraph(label, label_style), Paragraph(val_str, campo_style)])

            t_ficha = Table(ficha_data, colWidths=[5 * cm, 11.5 * cm])
            t_ficha.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, cor_leve]),
                ("BOX", (0, 0), (-1, -1), 0.5, cor_cinza),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, cor_cinza),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(t_ficha)

        story.append(Spacer(1, 1 * cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=cor_cinza))
        story.append(Spacer(1, 0.2 * cm))
        story.append(Paragraph(
            f"{INSTITUICAO}  ·  {UNIDADE_RESPONSAVEL}  ·  "
            f"Encarregada: {ENCARREGADA}  ·  Gerado em {data_geracao}",
            rodape_style,
        ))
        story.append(Paragraph(
            "Documento produzido nos termos da LGPD – Lei 13.709/2018, Art. 37 | "
            "Resolução TSE 23.222/2010 | Res. nº 971/2026",
            rodape_style,
        ))
        doc.build(story)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=relatorio_ropa_{ts}.pdf"},
        )

    flash("Formato inválido.", "danger")
    return redirect(url_for("index"))


@app.route("/seed", methods=["POST"])
@login_required
def seed():
    with get_conn() as conn:
        count = conn.execute("SELECT COUNT(*) FROM atividades WHERE ativo=1").fetchone()[0]
        if count > 0:
            flash("Base já possui registros. Seed não executado.", "warning")
            return redirect(url_for("index"))
        for ex in EXEMPLOS:
            conn.execute("""
                INSERT INTO atividades
                  (nome_atividade, finalidade, base_legal, categorias_titulares,
                   categorias_dados, dados_sensiveis, destinatarios, transferencia_inter,
                   prazo_retencao, medidas_seguranca, unidade_controladora, sistema_sei,
                   observacoes)
                VALUES
                  (:nome_atividade,:finalidade,:base_legal,:categorias_titulares,
                   :categorias_dados,:dados_sensiveis,:destinatarios,:transferencia_inter,
                   :prazo_retencao,:medidas_seguranca,:unidade_controladora,:sistema_sei,
                   :observacoes)
            """, ex)
    flash(f"{len(EXEMPLOS)} atividades de exemplo inseridas.", "success")
    return redirect(url_for("index"))


# ── Helper ────────────────────────────────────────────────────────────────────

def _form_to_dict(form) -> dict:
    return dict(
        nome_atividade=form.get("nome_atividade", "").strip(),
        finalidade=form.get("finalidade", "").strip(),
        base_legal=form.get("base_legal", "").strip(),
        categorias_titulares=form.get("categorias_titulares", "").strip(),
        categorias_dados=form.get("categorias_dados", "").strip(),
        dados_sensiveis=1 if form.get("dados_sensiveis") else 0,
        destinatarios=form.get("destinatarios", "").strip(),
        transferencia_inter=form.get("transferencia_inter", "N/A").strip(),
        prazo_retencao=form.get("prazo_retencao", "").strip(),
        medidas_seguranca=form.get("medidas_seguranca", "").strip(),
        unidade_controladora=form.get("unidade_controladora", "").strip(),
        sistema_sei=form.get("sistema_sei", "").strip(),
        observacoes=form.get("observacoes", "").strip(),
    )


# ── Entry point ───────────────────────────────────────────────────────────────

# Inicializa DB no import (cobre gunicorn e `python app.py`)
init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
